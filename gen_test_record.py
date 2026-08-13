import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- Styles ----
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_font = Font(name='Microsoft YaHei', size=10, bold=True)
title_font = Font(name='Microsoft YaHei', size=14, bold=True)
normal_font = Font(name='Microsoft YaHei', size=9)
small_font = Font(name='Microsoft YaHei', size=8)
red_bold_font = Font(name='Microsoft YaHei', size=9, bold=True, color='FF0000')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
light_blue_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
light_gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')

def apply_style(cell, font=normal_font, alignment=center_align, border=thin_border, fill=None):
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    if fill:
        cell.fill = fill

# Test conditions
test_points = [
    (1, '常温', 25, 37.0),
    (2, '常温', 25, 41.0),
    (3, '低温', 16, 37.0),
    (4, '低温', 16, 41.0),
    (5, '高温', 34, 37.0),
    (6, '高温', 34, 41.0),
]

headers = ['序号', '环温条件', '环温\n(℃)', '湿度\n(%RH)', '黑体设定\n(℃)', '黑体实测\n(℃)',
           '1#-第1次', '1#-第2次', '1#-第3次', '1#-均值', '1#-偏差',
           '2#-第1次', '2#-第2次', '2#-第3次', '2#-均值', '2#-偏差']

col_widths = [5, 10, 7, 7, 8, 8, 8, 8, 8, 8, 7, 8, 8, 8, 8, 7]

def create_mode_sheet(wb, sheet_name, title, header_fill, mode_label):
    ws = wb.create_sheet(sheet_name)

    # Title
    ws.merge_cells('A1:P1')
    ws['A1'] = title
    apply_style(ws['A1'], font=title_font)

    # Info
    ws.merge_cells('A2:P2')
    ws['A2'] = ('测试标准：目标温度35.0℃~42.0℃内 ±0.2℃；其他 ±0.3℃  |  '
                '测试仪器：水浴式黑体辐射源、高精度数字温度计  |  每点重复测量3次')
    apply_style(ws['A2'], font=small_font, alignment=left_align)

    # Headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        apply_style(cell, font=header_font, fill=header_fill)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    row = 4
    for idx, amb_cond, amb_t, bb_set in test_points:
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=amb_cond)
        ws.cell(row=row, column=3, value=amb_t)
        ws.cell(row=row, column=5, value=bb_set)
        for col in range(1, 17):
            cell = ws.cell(row=row, column=col)
            apply_style(cell, font=normal_font)
            if col in [4, 6, 7, 8, 9, 12, 13, 14]:
                cell.fill = yellow_fill
        row += 1

    # Mean & deviation formulas
    for r in range(4, 4 + 6):
        ws.cell(row=r, column=10).value = '=IF(COUNT(G%d:I%d)=3,AVERAGE(G%d:I%d),IF(COUNT(G%d:I%d)>0,AVERAGE(G%d:I%d),""))' % (r, r, r, r, r, r, r, r)
        apply_style(ws.cell(row=r, column=10), font=normal_font)
        ws.cell(row=r, column=11).value = '=IF(J%d="","",J%d-F%d)' % (r, r, r)
        apply_style(ws.cell(row=r, column=11), font=normal_font)
        ws.cell(row=r, column=15).value = '=IF(COUNT(L%d:N%d)=3,AVERAGE(L%d:N%d),IF(COUNT(L%d:N%d)>0,AVERAGE(L%d:N%d),""))' % (r, r, r, r, r, r, r, r)
        apply_style(ws.cell(row=r, column=15), font=normal_font)
        ws.cell(row=r, column=16).value = '=IF(O%d="","",O%d-F%d)' % (r, r, r)
        apply_style(ws.cell(row=r, column=16), font=normal_font)

    # Hint row
    ws.merge_cells('A%d:P%d' % (row, row))
    ws.cell(row=row, column=1, value='如需更多样品，可复制上方行并修改编号')
    apply_style(ws.cell(row=row, column=1), font=small_font, alignment=left_align, fill=light_gray_fill)
    row += 1

    # Criteria row
    ws.merge_cells('A%d:P%d' % (row, row))
    ws.cell(row=row, column=1, value='判定标准：37℃/41℃均在35~42℃范围内，偏差≤±0.2℃为合格（PASS），偏差>±0.2℃为不合格（FAIL）')
    apply_style(ws.cell(row=row, column=1), font=red_bold_font, alignment=left_align, fill=light_green_fill)
    row += 1

    # Signature
    ws.merge_cells('A%d:D%d' % (row, row))
    ws.cell(row=row, column=1, value='测试人：')
    apply_style(ws.cell(row=row, column=1), font=normal_font, alignment=left_align)
    ws.merge_cells('E%d:H%d' % (row, row))
    ws.cell(row=row, column=5, value='测试日期：')
    apply_style(ws.cell(row=row, column=5), font=normal_font, alignment=left_align)
    ws.merge_cells('I%d:L%d' % (row, row))
    ws.cell(row=row, column=9, value='审核人：')
    apply_style(ws.cell(row=row, column=9), font=normal_font, alignment=left_align)
    ws.merge_cells('M%d:P%d' % (row, row))
    ws.cell(row=row, column=13, value='审核日期：')
    apply_style(ws.cell(row=row, column=13), font=normal_font, alignment=left_align)

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 30

    return ws

# Delete default sheet
del wb['Sheet']

# Create 3 mode sheets
create_mode_sheet(wb, '耳温模式准确度', 'ET-2126 耳温模式准确度测试记录表', light_blue_fill, '耳温')
create_mode_sheet(wb, '额温模式准确度', 'ET-2126 额温模式准确度测试记录表', orange_fill, '额温')
create_mode_sheet(wb, '物温模式准确度', 'ET-2126 物温（Surface）模式准确度测试记录表', green_fill, '物温')

# ========== Sheet 4: 综合判定汇总 ==========
ws4 = wb.create_sheet('综合判定汇总')

ws4.merge_cells('A1:K1')
ws4['A1'] = 'ET-2126 准确度测试综合判定汇总表'
apply_style(ws4['A1'], font=title_font)

ws4.merge_cells('A2:K2')
ws4['A2'] = '汇总条件：环温25℃/16℃/34℃，水槽37℃/41℃，3种模式（耳温/额温/物温），每点3次测量取均值'
apply_style(ws4['A2'], font=small_font, alignment=left_align)

summary_headers = ['序号', '测试模式', '环温(℃)', '水槽温度(℃)', '样品编号',
                   '测量均值(℃)', '水槽实测(℃)', '偏差(℃)', '允许误差(℃)', '判定', '备注']
for col_idx, h in enumerate(summary_headers, 1):
    cell = ws4.cell(row=3, column=col_idx, value=h)
    apply_style(cell, font=header_font, fill=light_blue_fill)

summary_widths = [5, 10, 8, 10, 8, 10, 10, 8, 10, 8, 15]
for i, w in enumerate(summary_widths, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# 3 modes x 3 ambient x 2 bb x 2 samples = 36 rows
modes = ['耳温', '额温', '物温']
amb_temps = [25, 16, 34]
bb_temps = [37.0, 41.0]
samples = ['1#', '2#']

row = 4
idx = 1
for mode in modes:
    for amb in amb_temps:
        for bb in bb_temps:
            for sample in samples:
                ws4.cell(row=row, column=1, value=idx)
                ws4.cell(row=row, column=2, value=mode)
                ws4.cell(row=row, column=3, value=amb)
                ws4.cell(row=row, column=4, value=bb)
                ws4.cell(row=row, column=5, value=sample)
                ws4.cell(row=row, column=9, value='+/-0.2')
                for col in range(1, 12):
                    cell = ws4.cell(row=row, column=col)
                    apply_style(cell, font=normal_font)
                    if col in [6, 7, 8, 11]:
                        cell.fill = yellow_fill
                ws4.cell(row=row, column=10).value = '=IF(H%d="","",IF(ABS(H%d)<=0.2,"PASS","FAIL"))' % (row, row)
                apply_style(ws4.cell(row=row, column=10), font=normal_font)
                idx += 1
                row += 1

# Merge ambient cells for visual grouping
# Group by mode: each mode has 6 rows (3 amb x 2 bb x 2 samples)
for mode_start in range(4, 4 + 36, 12):
    for amb_group in range(0, 12, 4):
        start_r = mode_start + amb_group
        ws4.merge_cells('C%d:C%d' % (start_r, start_r + 3))

# Summary stats section
row += 2
ws4.merge_cells('A%d:K%d' % (row, row))
ws4.cell(row=row, column=1, value='判定统计')
apply_style(ws4.cell(row=row, column=1), font=Font(name='Microsoft YaHei', size=10, bold=True), fill=light_blue_fill)
row += 1

stats = [
    ('耳温 - PASS数', '=COUNTIFS(B4:B39,"耳温",J4:J39,"PASS")'),
    ('耳温 - FAIL数', '=COUNTIFS(B4:B39,"耳温",J4:J39,"FAIL")'),
    ('额温 - PASS数', '=COUNTIFS(B4:B39,"额温",J4:J39,"PASS")'),
    ('额温 - FAIL数', '=COUNTIFS(B4:B39,"额温",J4:J39,"FAIL")'),
    ('物温 - PASS数', '=COUNTIFS(B4:B39,"物温",J4:J39,"PASS")'),
    ('物温 - FAIL数', '=COUNTIFS(B4:B39,"物温",J4:J39,"FAIL")'),
]

for i, (label, formula) in enumerate(stats):
    r = row + i
    ws4.merge_cells('A%d:E%d' % (r, r))
    ws4.cell(row=r, column=1, value=label)
    apply_style(ws4.cell(row=r, column=1), font=normal_font, alignment=left_align)
    ws4.merge_cells('F%d:K%d' % (r, r))
    ws4.cell(row=r, column=6, value=formula)
    apply_style(ws4.cell(row=r, column=6), font=normal_font)

row += len(stats) + 1
ws4.merge_cells('A%d:E%d' % (row, row))
ws4.cell(row=row, column=1, value='总计 - PASS')
apply_style(ws4.cell(row=row, column=1), font=Font(name='Microsoft YaHei', size=10, bold=True), alignment=left_align)
ws4.merge_cells('F%d:K%d' % (row, row))
ws4.cell(row=row, column=6, value='=COUNTIF(J4:J39,"PASS")')
apply_style(ws4.cell(row=row, column=6), font=Font(name='Microsoft YaHei', size=10, bold=True))
row += 1
ws4.merge_cells('A%d:E%d' % (row, row))
ws4.cell(row=row, column=1, value='总计 - FAIL')
apply_style(ws4.cell(row=row, column=1), font=Font(name='Microsoft YaHei', size=10, bold=True, color='FF0000'), alignment=left_align)
ws4.merge_cells('F%d:K%d' % (row, row))
ws4.cell(row=row, column=6, value='=COUNTIF(J4:J39,"FAIL")')
apply_style(ws4.cell(row=row, column=6), font=Font(name='Microsoft YaHei', size=10, bold=True, color='FF0000'))

row += 2
ws4.merge_cells('A%d:D%d' % (row, row))
ws4.cell(row=row, column=1, value='测试人：')
apply_style(ws4.cell(row=row, column=1), font=normal_font, alignment=left_align)
ws4.merge_cells('E%d:H%d' % (row, row))
ws4.cell(row=row, column=5, value='测试日期：')
apply_style(ws4.cell(row=row, column=5), font=normal_font, alignment=left_align)
ws4.merge_cells('I%d:K%d' % (row, row))
ws4.cell(row=row, column=9, value='审核人：')
apply_style(ws4.cell(row=row, column=9), font=normal_font, alignment=left_align)

# Save
output_path = 'D:/考核项目/代码/10/ET-2126准确度测试记录表.xlsx'
wb.save(output_path)
print('Saved to:', output_path)
